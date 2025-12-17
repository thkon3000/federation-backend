# federation_app/excel_utils.py

from openpyxl import load_workbook
from datetime import datetime, date
from io import BytesIO


DATE_FORMATS = [
    "%d/%m/%Y",
    "%d-%m-%Y",
]


def parse_date(value):
    if value is None or value == "":
        return None

    if isinstance(value, date):
        return value

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, str):
        value = value.strip()
        for fmt in DATE_FORMATS:
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue

    return None


def read_members_excel(file_bytes: bytes):
    wb = load_workbook(BytesIO(file_bytes), data_only=True)

    # Try to find sheet named "ΜΕΛΗ" (robust)
    target_sheet = None
    for name in wb.sheetnames:
        if name.strip() == "ΜΕΛΗ":
            target_sheet = name
            break

    # Fallback: use first sheet
    if not target_sheet:
        target_sheet = wb.sheetnames[0]

    ws = wb[target_sheet]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip() if h else "" for h in rows[0]]
    data_rows = rows[1:]

    members = []

    for idx, row in enumerate(data_rows, start=2):
        if all(cell is None for cell in row):
            continue

        record = dict(zip(headers, row))
        record["_row"] = idx
        members.append(record)

    return members
